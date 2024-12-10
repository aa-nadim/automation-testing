import os
import logging
import pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook
from config.config import REPORT_FILE_PATH  # Ensure REPORT_FILE_PATH ends with ".xlsx"

# Ensure the reports directory exists
def create_report_directory():
    if not os.path.exists("reports"):
        os.makedirs("reports")
        logging.info("Created reports directory.")

def create_excel_file(filename=REPORT_FILE_PATH):
    """Create a new Excel file with a header if it doesn't exist."""
    if not os.path.isfile(filename):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Test Results"
        sheet.append(['Page URL', 'Test Case', 'Status', 'Comments'])
        workbook.save(filename)
        logging.info(f"Created new Excel report file: {filename}")

def read_existing_reports(filename=REPORT_FILE_PATH):
    """Read the existing Excel file and return a list of rows (test results)."""
    existing_rows = []
    if os.path.isfile(filename):
        workbook = load_workbook(filename)
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            existing_rows.append(row)
    return existing_rows

def write_to_excel(test_data, filename=REPORT_FILE_PATH):
    """Write test results to an Excel file."""
    create_report_directory()  # Ensure the reports directory exists
    create_excel_file(filename)  # Ensure the Excel file exists

    try:
        workbook = load_workbook(filename)
        sheet = workbook.active

        # Check if the test result already exists
        updated = False
        for row in sheet.iter_rows(min_row=2, values_only=True):  
            if row[0] == test_data[0] and row[1] == test_data[1]:
                sheet.cell(row=row[0].row, column=3, value=test_data[2])  
                sheet.cell(row=row[0].row, column=4, value=test_data[3])  
                updated = True
                break

        # If the test result didn't exist, append it
        if not updated:
            sheet.append(test_data)
            logging.info(f"Appended test result for {test_data[1]}.")

        workbook.save(filename)
        logging.info(f"Test results saved to {filename}.")

    except Exception as e:
        logging.error(f"Error writing to Excel: {e}")
        print(f"Error writing to Excel: {e}")

def run_test(driver, url, test_func, test_name):
    """Run a test function and write results to an Excel file."""
    try:
        logging.info(f"Running test: {test_name} on {url}")
        driver.get(url)
        
        # Log the start of the test
        print(f"Running {test_name} on {url}...")

        result, comments = test_func(driver)
        status = 'PASS' if result else 'FAIL'

        # Write results to Excel
        write_to_excel([url, test_name, status, comments])

        # Log the result of the test
        logging.info(f"Test result for {test_name}: {status}. Comments: {comments}")
        print(f"Test result for {test_name}: {status}. Comments: {comments}")
    except Exception as e:
        logging.error(f"Error during {test_name}: {e}")
        print(f"Error during {test_name}: {e}")
        write_to_excel([url, test_name, 'FAIL', str(e)])


def generate_report(test_results, test_name):
    """
    Generate an Excel report for a specific test with multiple sheets in one file.
    
    :param test_results: List of dictionaries containing test results
    :param test_name: Name of the test for report naming
    """
    # Create reports directory if not exists
    reports_dir = "reports"
    os.makedirs(reports_dir, exist_ok=True)

    # Create DataFrame from test results
    df = pd.DataFrame(test_results)

    # Generate unique filename with timestamp for the Excel file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{reports_dir}/test_result.xlsx"

    # Check if the file exists
    file_exists = os.path.exists(filename)
    
    # If the file exists, we append data to it
    if file_exists:
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
            df.to_excel(writer, index=False, sheet_name=test_name)
    else:
        # If the file doesn't exist, create it and write the data
        with pd.ExcelWriter(filename, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, index=False, sheet_name=test_name)

    print(f"Report generated: {filename}")
    return filename